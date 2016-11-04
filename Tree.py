#!/usr/bin/env python3

# external modules
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import BLUE


# my modules
import DCC
import Config as CF
import FileSys
import MyUtil

# Set Excel Styles
# Excel hyperlink style is calibri 11, underline blue
font_url_style = Font(color = BLUE, underline = 'single')
bold_style = Font(bold = True)
align_hv_cen_style = Alignment(horizontal = 'center', vertical = 'center')
align_ver_cen_style = Alignment(vertical = 'center')
align_hv_cen_wrap_style = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)

#function uses kwarg argument to check a load flag is given.  If a flag is given then the function will load the existing file. It gives the option to load from the disk or recreate the json file
def return_tree(s, target, rootfilename, **kwargs):
    load_flag = kwargs.get('Load')
        
    if FileSys.file_check_json(s, rootfilename) == True:
        if load_flag == None:
            print('File [',rootfilename,'] already exists - use existing file? ', sep = '', end = '')
            load_flag = MyUtil.get_yn('[Load from disk = Y, re-create = N] (Y/N)? ')
        if load_flag == True:       
            tr = FileSys.file_read_json(rootfilename)
            FileSys.file_write_json(tr, rootfilename, path = CF.dccfilepath)
            return(tr)
            
    tr = get_tree(s,target, **kwargs)
    FileSys.file_write_json(tr, rootfilename, path = CF.dccfilepath)
#     print_tree(s, tr)
    return(tr)

def flat_tree(tree, key, list):
    branch = tree[key]
    for doc in branch['documents']:
        list.append(doc)
    for other in branch['others']:
        list.append(other)
    for col in branch['collections']:
        list.append(col)
        flat_tree(tree, col, list)
    return(list)   

#function that iterates through the tree and prints document information
def iter_print_tree(s, tree, key, indent):
    branch = tree[key]
    for doc in branch['documents']:
        nameData = DCC.prop_get(s, doc, InfoSet = 'DocBasic')
        print(indent + doc)
        print(indent+'    DCC Title: ',nameData['title'],sep='') 
        print(indent+'    TMT Doc. Num.: ',nameData['tmtnum'],sep='') 
        print(indent+'    Owner: ',nameData['owner-name']) 
        print(indent+'    Filename: ',nameData['filename'],sep='')
        print(indent+'    Date Modified: ',nameData['date'],sep='')
        print(indent+'    URL: ','https://docushare.tmt.org/docushare/dsweb/ServicesLib/',doc,'/View',sep='')
    for other in branch['others']:
        nameData = DCC.prop_get(s, other, InfoSet = 'Title')
        print(indent+other, ':', nameData['title'])        
    for col in branch['collections']:
        nameData = DCC.prop_get(s, col, InfoSet = 'Title')
        print(indent+col, ':', nameData['title'])
        print(indent+'    URL: ','https://docushare.tmt.org/docushare/dsweb/ServicesLib/',col,sep='')        
        iter_print_tree(s, tree, col, indent+'    ')
# This function calls the iterative version of print tree
def print_tree(s,tree):
    iter_print_tree(s, tree, 'root', '')
# creates a hyperlink using the url given in arguments
def href_str(name, url):
    return('<a href="'+url+'">'+name+'</a>')
    
def url_view(handle):
    return('https://docushare.tmt.org/docushare/dsweb/ServicesLib/' + handle + '/View')
    
def url_access(handle):
    return('https://docushare.tmt.org/docushare/dsweb/View/' + handle)

def url_perm(handle):
    return('https://docushare.tmt.org/docushare/dsweb/ServicesLib/' + handle + '/Permissions')
    
def url_ver(handle):
    return('https://docushare.tmt.org/docushare/dsweb/ServicesLib/' + handle + '/History')
    
def url_loc(handle):
    return('https://docushare.tmt.org/docushare/dsweb/ServicesLib/' + handle + '/Location')

# function that iterates through html tree
def iter_html_tree(s, htmlfile, tree, key, indent, **kwargs):
    branch = tree[key]
    
    keyword = kwargs.get('Keyword', '')
    
    for doc in branch['documents']:
        nameData = DCC.prop_get(s, doc, InfoSet = 'DocBasic')
        if keyword in nameData['keywords']:
            print(doc)
            print('<div style="text-indent: ',str(indent),'em;">',file=htmlfile,sep='')
            print('<p>',file=htmlfile,sep='')
            print(href_str(nameData['title'],url_access(doc)),file=htmlfile,sep='')
            print('[',file=htmlfile,sep='')
            print(href_str(doc,url_view(doc)),file=htmlfile,sep='') 
            print(', ',file=htmlfile,sep='')
            print(href_str('Perm',url_perm(doc)),file=htmlfile,sep='')
            print(', ',file=htmlfile,sep='')
            print(href_str('Ver',url_ver(doc)),file=htmlfile,sep='')
            print(', ',file=htmlfile,sep='')
            print(href_str('Loc',url_loc(doc)),file=htmlfile,sep='')
            print(']',file=htmlfile,sep='')
            print('</p>',file=htmlfile,sep='')
            print('</div>',file=htmlfile,sep='') 
                
            print('<div style="text-indent: ',str(indent+2),'em;">',file=htmlfile,sep='')
            print('<p>TMT Doc. Num.: ',nameData['tmtnum'],'</p>',file=htmlfile,sep='') 
            print('<p>Owner: ',nameData['owner-name'],file=htmlfile,sep='')
            print('<p>Filename: ',nameData['filename'],file=htmlfile,sep='')
            print('<p>Date Modified: ',nameData['date'],file=htmlfile,sep='')
            print('</div>',file=htmlfile,sep='') 
                
    for other in branch['others']:
        print(other)
        nameData = DCC.prop_get(s, other, InfoSet = 'Title')
        print('<div style="text-indent: ',str(indent),'em;">',file=htmlfile,sep='')
        print('<p>',file=htmlfile,sep='')
        print('[',file=htmlfile,sep='')
        print(href_str(other,url_access(other)),file=htmlfile,sep='') 
        print(']',file=htmlfile,sep='')
        print(href_str(nameData['title'],url_view(other)),file=htmlfile,sep='')
        print('[',file=htmlfile,sep='')
        print(href_str('Perm',url_perm(other)),file=htmlfile,sep='')
        print(']',file=htmlfile,sep='')

        print('</p>',file=htmlfile,sep='')
        print('</div>',file=htmlfile,sep='')        
        
    for col in branch['collections']:
        print(col)
        nameData = DCC.prop_get(s, col, InfoSet = 'Title')
        print('<div style="text-indent: ',str(indent),'em;">',file=htmlfile,sep='')
        print('<p></p>',file=htmlfile,sep='')
        print('<p>',file=htmlfile,sep='')
        print(href_str(nameData['title'],url_access(col)),file=htmlfile,sep='')
        print('[',file=htmlfile,sep='')
        print(href_str(col,url_view(col)),file=htmlfile,sep='') 
        print(', ',file=htmlfile,sep='')
        print(href_str('Perm',url_perm(col)),file=htmlfile,sep='')
        print(', ',file=htmlfile,sep='')
        print(href_str('Ver',url_ver(col)),file=htmlfile,sep='')
        print(', ',file=htmlfile,sep='')
        print(href_str('Loc',url_loc(col)),file=htmlfile,sep='')
        print(']',file=htmlfile,sep='')

        print('</p>',file=htmlfile,sep='')
        print('</div>',file=htmlfile,sep='')        
        iter_html_tree(s, htmlfile, tree, col, indent+2, **kwargs)
        
def html_tree(s,tree,froot,**kwargs):
    htmlfile = open(CF.reportfilepath + froot+'.html','w+')
    print('<!DOCTYPE html><html><body>',file = htmlfile)
    print('<style> p {line-height: 0.35} </style>', file = htmlfile)
    print('<h1>',froot,'</h1>', file = htmlfile)

    iter_html_tree(s, htmlfile, tree, 'root', 0, **kwargs)
    
    print('</body></html>',file = htmlfile)  
    htmlfile.close
    
# function sets excel workbook headings for html tree
def xls_tree_headings(ws):
    col = 1
    ws.cell(row = 1, column = col).value = "ID"
    col += 1
    ws.cell(row = 1, column = col).value = "Collection Handle"
    col += 1
    ws.cell(row = 1, column = col).value = "Collection Title"
    col += 1
    ws.cell(row = 1, column = col).value = "Document Title"
    col += 1
    ws.cell(row = 1, column = col).value = "TMT Doc Number"
    col += 1    
    ws.cell(row = 1, column = col).value = "Document Handle"
    col += 1    
    ws.cell(row = 1, column = col).value = "Version Handle"   
    col += 1
    ws.cell(row = 1, column = col).value = "Owner"
    col += 1
    ws.cell(row = 1, column = col).value = "File Name"
    col += 1
    ws.cell(row = 1, column = col).value = "Date Modified"
    
    colcnt = col+1
    # sets alignment and font style for headings
    for col in range(1, colcnt):
        ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
        ws.cell(row = 1, column = col).font = bold_style
        
    # Set column widths
    ws.column_dimensions["A"].width = 5.0
    ws.column_dimensions["B"].width = 15.0
    ws.column_dimensions["C"].width = 25.0
    ws.column_dimensions["D"].width = 60.0
    ws.column_dimensions["E"].width = 25.0
    ws.column_dimensions["F"].width = 15.0    
    ws.column_dimensions["G"].width = 15.0
    ws.column_dimensions["H"].width = 20.0
    ws.column_dimensions["I"].width = 60.0
    ws.column_dimensions["J"].width = 35.0

# Prints data to xls document
def xls_print_ssrow(ws, collData, docData, ssrow):
    col = 1
    
    # Column 1: ID    
    ws.cell(row = ssrow, column = col).value = ssrow-1
    ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style
    col += 1
    
    # Column 2: Coll Handle
    ws.cell(row = ssrow, column = col).value = collData['handle']
    ws.cell(row = ssrow, column = col).font = font_url_style
    ws.cell(row = ssrow, column = col).hyperlink = url_view(collData['handle'])
    ws.cell(row = ssrow, column = col).alignment = Alignment(vertical = 'center')
    col += 1

    # Column 3: Coll Name
    ws.cell(row = ssrow, column = col).value = collData['title']
    ws.cell(row = ssrow, column = col).font = font_url_style
    ws.cell(row = ssrow, column = col).hyperlink = url_access(collData['handle'])
    ws.cell(row = ssrow, column = col).alignment = Alignment(vertical = 'center')
    col += 1    
    
    # Column 4: Doc Title
    ws.cell(row = ssrow, column = col).value = docData['title']
    ws.cell(row = ssrow, column = col).font = font_url_style
    ws.cell(row = ssrow, column = col).hyperlink = url_access(docData['handle'])
    ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
    col += 1   

    # Column 5: TMT Document Number
    ws.cell(row = ssrow, column = col).value = docData['tmtnum']
    ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
    col += 1  
    
    # Column 6: Doc Handle
    ws.cell(row = ssrow, column = col).value = docData['handle']
    ws.cell(row = ssrow, column = col).font = font_url_style
    ws.cell(row = ssrow, column = col).hyperlink = url_view(docData['handle'])
    ws.cell(row = ssrow, column = col).alignment = Alignment(vertical = 'center')
    col += 1  
    
    # Column 7: Ver Handle
    ws.cell(row = ssrow, column = col).value = docData['Versions']['prefver']
    ws.cell(row = ssrow, column = col).font = font_url_style
    ws.cell(row = ssrow, column = col).hyperlink = url_view(docData['Versions']['prefver'])
    ws.cell(row = ssrow, column = col).alignment = Alignment(vertical = 'center')
    col += 1   

    # Column 8: Owner
    ws.cell(row = ssrow, column = col).value = docData['owner-name']
    ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style
    col += 1   

    # Column 9: File Name
    ws.cell(row = ssrow, column = col).value = docData['filename']
    ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
    col += 1   
    
    # Column 10: Date Modified
    ws.cell(row = ssrow, column = col).value = docData['date']
    ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style

# Global Variable ssrow
ssrow = 2

#iterates through xls version of html tree
def xls_tree_iter(s,ws,tree,col, **kwargs):
    global ssrow
    # Write and format the headings in Excel
    xls_tree_headings(ws)

    collData = DCC.prop_get(s, col, InfoSet = 'Title')
    branch = tree[col]
    
    keyword = kwargs.get('Keyword', '')
    
    for doc in branch['documents']:
        print(col,doc)
        docData = DCC.prop_get(s, doc, InfoSet = 'DocBasic')
        docData['Versions'] = DCC.prop_get(s,doc,InfoSet = 'Versions',WriteProp = True)
        if keyword in docData['keywords']:
            xls_print_ssrow(ws, collData, docData, ssrow)
            ssrow += 1
    for newcol in branch['collections']:
        xls_tree_iter(s,ws,tree,newcol,**kwargs)
        
def xls_tree(s,tree,col,fname, **kwargs):
    # Open the spreadsheet
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    
    xls_tree_iter(s,ws,tree,col,**kwargs)
    
    wb.save(CF.reportfilepath + fname + '.xls')   

# Funciton builds html tree and uses prop_get to get information of the target collection
# Uses kwarg argument to recognize exclude list
def build_tree(s, keyname, target, tree, **kwargs):
    # kwargs options:
    #  Exclude - List of handles to not be included in the tree
    
    excludeList = kwargs.get('Exclude',[])

    documents = []
    collections = []
    others = []
    dict = {}
    
    fd = DCC.prop_get(s, target, InfoSet = 'CollCont', Depth = '1')

    for idx,d in enumerate(fd):
        handle = d['name'][1]
        print(handle)
        if not handle in excludeList:
            if idx == 0:
                dict['parent'] = handle
            else:
                if 'Document' in handle:
                    documents.append(handle)
                elif 'Collection' in handle:
                    collections.append(handle)
                else:
                    others.append(handle)

    dict['collections'] = collections
    dict['documents'] = documents
    dict['others'] = others

    tree[keyname] = dict
    for col in collections:
        if not col in excludeList:
            tree = build_tree(s, col, col, tree, **kwargs)
    return(tree)

def get_tree(s, collhandle, **kwargs):
    return(build_tree(s, collhandle, collhandle, build_root(collhandle), **kwargs))
    
def get_flat_tree(tree):
    fl = flat_tree(tree, 'root', [])    
    return(fl)
    
def build_root(collhandle):
    tree = {}
    tree['root'] = {'collections' : [collhandle], 'documents' : [], 'others' : []}
    return(tree)

def test_tree():
    collhandle = 'Collection-286'
    exclude = ['Collection-7337','Document-21244', 'Document-26018']

    # Login to DCC
    s = DCC.login(CF.dcc_url + CF.dcc_login)

    print('excluding:',exclude)
    tree = get_tree(s, collhandle, Exclude = exclude)
    print_tree(s,tree)
    
    print('\n\n')
    for branch in tree:
        print(branch+': ',tree[branch])
        
    fl = flat_tree(tree, 'root', [])
    print(fl)
    
if __name__ == '__main__':
    print("Running module test code for",__file__)
#     test_tree()
    # Login to DCC
    s = DCC.login(CF.dcc_url + CF.dcc_login) 

#     froot = 'Listing of Scott Collection'
#     coll = 'Collection-286'
    
#     froot = 'Listing M1CS Preliminary Design Review Collections'
#     coll = 'Collection-10725'
#     
#     tr = return_tree(s, coll, froot)
# #     print_tree(s,tr)
#     html_tree(s,tr,froot)
    
#     load_flag = True

#     froot = 'Listing of WFOS for Suijian'
#     coll = 'Collection-7798'
#     tr = return_tree(s, coll, froot)
#     html_tree(s,tr,froot)
 
    froot = 'Test of HTML bug'
    coll = 'Collection-2656'
    tr = return_tree(s, coll, froot)
    html_tree(s,tr,froot)   

#     froot = 'Listing of IRIS'
#     coll = 'Collection-2463'

#     froot = 'Listing of STR CID'
#     coll = 'Collection-10669'
#     coll = 'Collection-11377'
#     tr = return_tree(s, coll, froot)
#     xls_tree(s,tr,coll,froot)
#     html_tree(s,tr,froot)
    
#     froot = 'Listing of STR CID'
#     coll = 'Collection-10668'
#     tr = return_tree(s, coll, froot)
#     html_tree(s,tr,froot)
    
#     froot = 'Listing of 2013 [Nov 12-14] Telescope Structure System Preliminary Design Review'
#     coll = 'Collection-7095'
#     tr = return_tree(s, coll, froot, Load = load_flag)
#     html_tree(s,tr,froot)
#     
#     froot = 'Listing of 2014 [April 15 -16] Telescope Structure System Controls-Software Preliminary Design Review'
#     coll = 'Collection-7857'
#     tr = return_tree(s, coll, froot, Load = load_flag)
#     html_tree(s,tr,froot)
# 
#     froot = 'Listing of 2014 [April 15 -16] Telescope Structure System Controls-Software Preliminary Design Review'
#     coll = 'Collection-7857'
#     tr = return_tree(s, coll, froot, Load = load_flag)
#     html_tree(s,tr,froot)
#     
#     froot = 'Listing of 2014 [Nov 18-20] Segment Handling System Preliminary Design Review'
#     coll = 'Collection-8674'
#     tr = return_tree(s, coll, froot, Load = load_flag)
#     html_tree(s,tr,froot)
#     
#     froot = 'Listing of 2015 [Feb 17-20] Telescope Structure (STR) Final Design Review Presentation (FDRP1)'
#     coll = 'Collection-9067'
#     tr = return_tree(s, coll, froot, Load = load_flag)
#     html_tree(s,tr,froot)
#     
#     froot = 'Listing of 2015 [Jul 27-29] Telescope Structure (STR) Final Design Review Presentation (FDRP2)'
#     coll = 'Collection-10071'
#     tr = return_tree(s, coll, froot, Load = load_flag)
#     html_tree(s,tr,froot)
#     
#     froot = 'Listing of 2015 [Oct 08-09] Telescope Structure (STR) Long Lead Procurement Review (LPR)'
#     coll = 'Collection-11045'
#     tr = return_tree(s, coll, froot, Load = load_flag)
#     html_tree(s,tr,froot)
    
    
