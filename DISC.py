#!/usr/bin/env python3

# reference: http://www.crummy.com/software/BeautifulSoup/bs4/doc/

# External modules
import requests
import string
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Style, Alignment
from openpyxl.styles.colors import BLUE
from openpyxl import load_workbook
from datetime import datetime

# My modules
import Config as CF
import DCC

# Set Excel Styles
# Excel hyperlink style is calibri 11, underline blue
font_url_style = Font(color = BLUE, underline = 'single')
bold_style = Font(bold = True)
align_hv_cen_style = Alignment(horizontal = 'center', vertical = 'center')
align_ver_cen_style = Alignment(vertical = 'center')

def get_bb_prop(s, handle):
# Get properties from a bulletin board entry
# Inputs:
#     s - requests session object
#     handle - bb identifier string of the form 'bulletin-XXXX'
# Returns:
#     list containing title, description and keywords strings


    url = CF.dcc_url + "/dsweb/PROPFIND/" + handle
    headers = {"DocuShare-Version":"5.0", "Content-Type":"text/xml", "Accept":"text/xml","Depth":"0"}
    xml = """<?xml version="1.0" ?>
        <propfind>
          <prop>
            <title/><description/><keywords/>
          </prop>
        </propfind>"""
    r = s.post(url,data=xml,headers=headers) 
    r.raise_for_status()
#     print('Status code:', r.status_code) 
    
    # Retrieve the results
    dom = BeautifulSoup(r.text)
    title = dom.title.text.strip()
    keywords = dom.keywords.text.strip()
    
    # Have to further process the description
    dom = BeautifulSoup(dom.description.text)
    description = ''
    for string in dom.strings:
        description = description.strip() + ' ' + string.strip()
    description = clean_string(description.strip())

#     print('title:', title)
#     print('description:', description)
#     print('keywords:', keywords)
    return([title, description, keywords])

def set_bb_prop(s, title, description, keywords, handle):
    url = CF.dcc_url + "/dsweb/PROPPATCH/" + handle
    headers = {"DocuShare-Version":"5.0", "Content-Type":"text/xml", "Accept":"text/xml"}
    xml = """<?xml version="1.0" ?>
        <propertyupdate>
          <set><prop>
            <title><![CDATA[""" + title  + """]]></title>
            <description><![CDATA[""" + description  + """]]></description>
            <keywords><![CDATA[""" + keywords  + """]]></keywords>
          </prop></set>
        </propertyupdate>"""
    r = s.post(url,data=xml,headers=headers) 
    r.raise_for_status()
#     print('Status code:', r.status_code)
    
def create_bb_post(s, title, description, keywords, handle):
    url = CF.dcc_url + "/dsweb/MKRES/" + handle + "/Bulletin"
    headers = {"DocuShare-Version":"5.0", "Content-Type":"text/xml", "Accept":"text/xml"}
    xml = """<?xml version="1.0" ?>
        <propertyupdate>
          <set><prop>
            <title><![CDATA[""" + title  + """]]></title>
            <description><![CDATA[""" + description  + """]]></description>
            <keywords><![CDATA[""" + keywords  + """]]></keywords>
          </prop></set>
        </propertyupdate>"""
    r = s.post(url,data=xml,headers=headers) 
    r.raise_for_status()
#     print('Status code:', r.status_code)

def clean_string(str):
    # Split text on newlines and add them to the list
    # Reference: http://stackoverflow.com/questions/29976234/openpyxl-unicode-values
    try:
        outstr = ''
        for str in str.replace('_x000D_','').splitlines():
            outstr = outstr + ' ' + str
    except:
        outstr = ''
        print('clean_str: exception encountered for string', str)

    outstr = outstr.strip()
    return(outstr)


def get_postentry_author_info( entry ):
    pa = entry.find("td", class_= "postentry_author")
    author = pa("input")[0].attrs['value']
    auth_url = 'https://docushare.tmt.org/' + pa("a")[0].attrs['href']

    # The above author = line is the same as:
    #   td_author = postentry.find("td", class_="postentry_author")
    #   input_author = td_author("input")
    #   author = input_author[0].attrs['value']
    
    # Find the date and time of the original entry
    dt = pa("span")[0]("span")[0].string.strip()

    # Format it as a string
    # time format: 02/11/15 11:18 AM
    dt = datetime.strptime(dt, "%m/%d/%y %I:%M %p")
    
    ent = entry.find("div", class_ = "postentry_header").find("a")
    url = 'https://docushare.tmt.org/' + ent.attrs["href"]
    title = ent.text
    
    # If there are carriage returns in the entry there will be multiple "p" instances.
    try:
        ent_str = entry.find("div", class_ = "postdescription").text.strip()
    except:
        pentries = entry.find("div", class_ = "postdescription").find_all("p")
        ent_str = ''
        for ent in pentries:
            ent_str = ent_str + '\t' + ent.text
    
    return([title, url, author, auth_url, dt, ent_str])

def set_ss_headings(ws, rownum, col):
    
    ws.cell(row = rownum, column = col).value = "ID"
    ws.cell(row = rownum, column = col).alignment = align_hv_cen_style 

    col += 1
    ws.cell(row = rownum, column = col).value = "Title"
    ws.cell(row = rownum, column = col).alignment = align_ver_cen_style 

    col += 1    
    ws.cell(row = rownum, column = col).value = "Posting"
    ws.cell(row = rownum, column = col).alignment = align_ver_cen_style 

    col += 1    
    ws.cell(row = rownum, column = col).value = "Author"
    ws.cell(row = rownum, column = col).alignment = align_ver_cen_style

    col += 1    
    ws.cell(row = rownum, column = col).value = "Post Date"
    ws.cell(row = rownum, column = col).alignment = align_hv_cen_style 

    col += 1
    ws.cell(row = rownum, column = col).value = "# Replies"
    ws.cell(row = rownum, column = col).alignment = align_hv_cen_style 

    col += 1    
    ws.cell(row = rownum, column = col).value = "Replies By:"
    ws.cell(row = rownum, column = col).alignment = align_hv_cen_style

    col += 1
    ws.cell(row = rownum, column = col).value = "Latest Reply Date"
    ws.cell(row = rownum, column = col).alignment = align_hv_cen_style

    col += 1    	
    ws.cell (row = rownum, column = col).value = "Latest Reply"
    ws.cell (row = rownum, column = col).alignment = align_hv_cen_style
    
    col += 1    	
    ws.cell (row = rownum, column = col).value = "Disposition"
    ws.cell (row = rownum, column = col).alignment = align_hv_cen_style
    
    col += 1    	
    ws.cell (row = rownum, column = col).value = "Action"
    ws.cell (row = rownum, column = col).alignment = align_hv_cen_style
	
    colmax = col + 1
    for col in range(1, colmax):
        ws.cell(row = rownum, column = col).font = bold_style
        
def get_discussion_rowcol(url, htmlfile, xlfile, ssrow, sscol):
    # Login to DCC and save the discussion as an html file
    s = DCC.login(CF.dcc_url + CF.dcc_login)
    res = s.get(url)
    res.raise_for_status()

    webfile = open(CF.dccfilepath + htmlfile,'wb')
    for chunk in res.iter_content(100000):
        webfile.write(chunk)
    webfile.close

    # Get the HTML into the Beautiful Soup object
    dcc=open(CF.dccfilepath + htmlfile,'r',encoding='utf-8').read()
    dom = BeautifulSoup(dcc, "html.parser")

    # Open the spreadsheet
    try:
        wb = load_workbook(xlfile)
        print('Opened existing file :', xlfile)
        existing_doc = True
    except:
        wb = openpyxl.Workbook()
        print('Created new file :', xlfile)
        existing_doc = False
        
    ws = wb.worksheets[0]
    
    # Write and format the headings in Excel if this is a new document
    set_ss_headings(ws, ssrow, sscol)
    

    # Find the Parent of all the original posts (that may have replies)
    # This is <form name="ToolbarMulti" method="post" action="/docushare/dsweb/ProcessMultipleCommand">

    form_tag = dom.find("form", {"name":"ToolbarMulti"})       

    rowoff = ssrow
    ssrow = ssrow + 1

    # Now find all the children that are Post Entries

    for idx, postentry in enumerate(form_tag.find_all("div", class_ = "postentry", recursive = False)):

        [title, url, author, auth_url, dt, post] = get_postentry_author_info( postentry )

        # post is the text of the original posting.
        # Look for replies to posts
        # The replyposts class is always the next sibling

        replies = postentry.find_next_sibling()

        # count the number of replies
        times = 0    
        reps = []
        r_dates = [dt]
        r_latest = ''
        r_title = ''
        r_disposition = ''
        r_action = ''

        for replyentry in replies.find_all("div", class_ = "postentry"):
            [r_title, r_url, r_author, r_auth_url, r_dt, r_latest] = get_postentry_author_info( replyentry )

            # Put results into a list with line breaks
            reps.append(r_author + ': ' + r_dt.strftime('%y-%m-%d %H:%M'))

            r_dates.append(r_dt)  
            
            if '_DISPOSITION' in r_title.upper():
#                 print('r_disposition is :', r_latest)
                r_disposition = r_latest
                                
            if '_ACTION' in r_title.upper():
#                 print('r_action is :', r_latest)
                r_action = r_latest
          
            times += 1

        # Turn the reps list into a single string    
        repstr = ';\n'.join(reps) 

        # Find the latest modified date
        r_dates.sort()
        r_date_latest = r_dates[-1]
        r_date_latest_str = r_date_latest.strftime('%y-%m-%d %H:%M')

        # Print output
        print('\nEntry:', idx)
        print('title:', title) 
        print('url:', url)
        print('posting:', post)
        print('author:', author)
        print('date:',dt.strftime('%y-%m-%d %H:%M'))
        if repstr != '':
            print('latest update:', r_date_latest_str)
            print('replies by:', repstr)
            print('latest reply title:', r_title)
            print('latest reply:', r_latest)
        if r_disposition != '':
            print('Disposition:',r_disposition)
        if r_action != '':
            print('Action:', r_action)

        # Column 1: ID
        col = sscol
        ws.cell(row = ssrow, column = col).value = ssrow-rowoff
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style

        # Column 2: Title
        col += 1
        ws.cell(row = ssrow, column = col).value = title
        ws.cell(row = ssrow, column = col).font = font_url_style
        ws.cell(row = ssrow, column = col).hyperlink = url
        ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')

        # Column 3: Posting
        col += 1
        ws.cell(row = ssrow, column = col).value = clean_string(post)                     
        ws.cell(row = ssrow, column = col).alignment =  Alignment(wrap_text = True, vertical = 'center')
        
        # Column 4: Author
        col += 1
        ws.cell(row = ssrow, column = col).value = author
        ws.cell(row = ssrow, column = col).hyperlink = auth_url
        ws.cell(row = ssrow, column = col).alignment = align_ver_cen_style
        ws.cell(row = ssrow, column = col).font = font_url_style

        # Column 5: Post Date
        col += 1
        ws.cell(row = ssrow, column = col).value = dt
        ws.cell(row = ssrow, column = col).number_format = 'YY-MM-DD HH:MM' 
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style   

        # Column 6: # Replies
        col += 1
        ws.cell(row = ssrow, column = col).value = times  
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style 

        # Column 7: Replies By
        col += 1
        ws.cell(row = ssrow, column = col).value = repstr
        ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')

        # Column 8: Latest Reply Date
        col += 1
        ws.cell(row = ssrow, column = col).value = r_date_latest_str
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style  
		
		#Column 9: Latest Reply Text
        col += 1
        ws.cell(row = ssrow, column = col).value = clean_string(r_latest)
        ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
        
        #Column 10: Disposition Text
        col += 1
        ws.cell(row = ssrow, column = col).value = clean_string(r_disposition)
        ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
        
        #Column 11: Action Text
        col += 1
        ws.cell(row = ssrow, column = col).value = clean_string(r_action)
        ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
        
        ssrow += 1

    # Set column widths
    ws.column_dimensions["A"].width = 5.0
    ws.column_dimensions["B"].width = 40.0
    ws.column_dimensions["C"].width = 60.0
    ws.column_dimensions["D"].width = 13.0
    ws.column_dimensions["E"].width = 13.0
    ws.column_dimensions["F"].width = 13.0    
    ws.column_dimensions["G"].width = 25.0
    ws.column_dimensions["H"].width = 17.0
    ws.column_dimensions["I"].width = 60.0
    ws.column_dimensions["J"].width = 60.0
    ws.column_dimensions["K"].width = 60.0

    # Save the spreadsheet
    wb.save(xlfile)


def get_discussion(url, htmlfile, xlfile):
    get_discussion_rowcol(url, htmlfile, xlfile, 1, 1)
    
def test_bulletin_post():
    # Login to DCC
    s = DCC.login(CF.dcc_url + CF.dcc_login)
    
    handle = 'Bulletin-6185'
    
    title = '_DISPOSITION'
    
    description = 'This item is open. MELCO made no reply.'
#     description = 'This can be closed with a tracked action'
#     description = "The reply doesn't answer directly to the question of the RIX. The reply seems to say that DP04S-A isn't a document to be reviewed. Is DP04S-A  a review item of FDRP2? Need more clear reply."
    
    
    keywords = 'Reviewer Disposition'
    
    create_bb_post(s, title, description, keywords, handle)
    
    title = '_ACTION'
    description = 'ACTION: TBD (To Be Determined) after issue is addressed by MELCO.'
#     description = 'ACTION: MELCO to study what is the probability of the scenario.'
#     description = "ACTION: MELCO to add all the safety related drives to the Table 2.1-1."
    keywords = 'Reviewer ACTION'
    
    create_bb_post(s, title, description, keywords, handle)

# if __name__ == '__main__':
#     print("Running module test code for",__file__)
#     test_bulletin_post()

